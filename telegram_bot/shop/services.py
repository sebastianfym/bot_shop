import os
from django.conf import settings
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook


def filling_excel(order_data, total_price, goods):

    file_path = os.path.join(settings.BASE_DIR, 'shop', 'file', 'order_file.xlsx')
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
        row_number = sheet.max_row + 1
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Адрес доставки', 'Имя получателя', 'Телефон получателя', 'Товары', 'Сумма заказа'])
        row_number = 2

    data_to_add = [order_data['delivery_address'], order_data['recipient_name'], order_data['recipient_phone'], goods,
                   f"{total_price} руб"]
    sheet.append(data_to_add)

    workbook.save(file_path)